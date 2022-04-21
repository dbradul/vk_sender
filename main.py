import os
import random
import sys
import time

import requests
import vk_api
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from vk_common.models import VkClientProxy
from vk_common.utils import logger, login_retrier, repack_exc

load_dotenv()

VIDEO_ENDINGS = ['.mp4', '.webm', '.mkv', '.mov', '.avi', '.flv', '.ogg', '.wmv']
NUM_ACCOUNTS_THRESHOLD = int(os.getenv('NUM_ACCOUNTS_THRESHOLD'))
NUM_CALLS_THRESHOLD = int(os.getenv('NUM_CALLS_THRESHOLD'))
MIN_WAIT = int(os.getenv('MIN_WAIT'))
MAX_WAIT = int(os.getenv('MAX_WAIT'))

@login_retrier
@repack_exc
# @login_enforcer(num_calls_threshold=int(os.getenv('NUM_CALLS_THRESHOLD')))
def send_message_video(client: VkClientProxy, user_id, message, mediaurl):
    vk_uploader = vk_api.upload.VkUpload(client._obj)

    res_upload_video = vk_uploader.video(
        video_file=mediaurl,
        is_private=True
    )

    media_file_id = f'video{res_upload_video["owner_id"]}_{res_upload_video["video_id"]}'

    res_send = client.messages.send(
        user_id=user_id,
        random_id=random.randint(100, 100000000),
        message=message,
        attachment=media_file_id
    )

    return res_send


@login_retrier
@repack_exc
# @login_enforcer(num_calls_threshold=int(os.getenv('NUM_CALLS_THRESHOLD')))
def send_message_photo(client, user_id, message, mediaurl):

    album_id = 0
    res_get_albums = client.photos.getAlbums(
        # owner_id=client.owner_id
    )

    if res_get_albums.get('items'):
        album_id = res_get_albums['items'][0]['id']
        client.owner_id = res_get_albums['items'][0]['owner_id']

    if not album_id:
        res_create_album = client.photos.createAlbum(
            title='main',
        )
        album_id = res_create_album['id']
        client.owner_id = res_create_album['owner_id']

    res_upload = client.photos.getUploadServer(
        album_id=album_id
    )

    with open(mediaurl, 'rb') as f:
        res = requests.post(
            url=res_upload['upload_url'],
            files={
                'photo': f
            }
        )

    json_res = res.json()

    res_save = client.photos.save(
        server=json_res['server'],
        photos_list=json_res['photos_list'],
        aid=json_res['aid'],
        hash=json_res['hash'],
        album_id=album_id
    )

    media_file_id = f'photo{client.owner_id}_{res_save[0]["id"]}'

    res_send = client.messages.send(
        user_id=user_id,
        random_id=random.randint(100, 100000000),
        message=message,
        attachment=media_file_id
    )

    return res_send


@login_retrier
@repack_exc
def get_user_id_by_name(client, user_name): # TODO: replace with resolve method
    users = client.users.get(
        user_ids=[user_name]
    )
    if not users:
        raise RuntimeError(f'User not found: {user_name}')

    return users[0]['id']


def process_file(client, filename):
    wb = load_workbook(filename)
    ws: Worksheet = wb.active
    sheet = wb.worksheets[0]
    row_count = sheet.max_row

    for idx, r in enumerate(range(2, row_count + 1), 1):
        try:
            vkurl = ws[f'F{str(r)}'].value
            message = ws[f'G{str(r)}'].value
            mediaurl = ws[f'H{str(r)}'].value
            user_name = vkurl.split('/')[-1]
            user_id = get_user_id_by_name(client, user_name)
            if any([True for elem in VIDEO_ENDINGS if mediaurl.lower().endswith(elem)]):
                resp = send_message_video(client, user_id=user_id, message=message, mediaurl=mediaurl)
            else:
                resp = send_message_photo(client, user_id=user_id, message=message, mediaurl=mediaurl)
            ws[f'J{str(r)}'].value = resp
            logger.info(f'Processed {vkurl} with result {resp}')
            wb.save(filename)
            time.sleep(random.randint(MIN_WAIT, MAX_WAIT))
        except Exception as ex:
            logger.error(f'Failed to send message for user {vkurl}: {ex}')

    wb.save(filename)


def main():

    vk_client = VkClientProxy(
        num_calls_threshold=NUM_CALLS_THRESHOLD,
        num_accounts_threshold=NUM_ACCOUNTS_THRESHOLD,
        call_domain='messages'
    )
    vk_client.load_accounts()
    vk_client.auth_until_success()

    if len(sys.argv) > 1:
        param = sys.argv[1]
        process_file(vk_client, param)


if __name__ == '__main__':
    main()

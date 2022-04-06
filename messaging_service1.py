import json
import logging
import random
import requests
from openpyxl.worksheet.worksheet import Worksheet
from requests import Response
from openpyxl import Workbook, load_workbook
import time

logger = logging.getLogger(__name__)

TARGET_URL = 'https://vk.com/id711959514'
BASE_URL = 'https://api.vk.com/'
AUTH_URL = 'https://oauth.vk.com/token?grant_type=password&client_id=2274003&client_secret=hHbZxrka2uZ6jB1inYsH&'
SEND_MESSAGE_PATH = 'method/messages.send'
GET_USER_PATH = 'method/users.get'

A = 'Номер'
B = 'Імя'
C = 'Прізвище'
D = 'По батькові'
E = 'Народження'
F = 'VKURL'
G = 'Message'
H = 'MediaURL'
J = 'STATUS'


class MessagingService:
    """
    VK REST API service (https://dev.vk.com/reference)
    Currently only send message is implemented, but any other method can be implemented in the same way.

    Example of usage:
    '''
        ms = MessagingService(username='+380630597236', password='Ljn%26f%232!')
        user_id = ms.url_to_id(TARGET_URL)
        r = ms.send_message(user_id, '111')
    '''
    """

    def __init__(self, username, password, excel_path, timeout: int):
        self.token, self.user_id = self.get_token(username, password)
        self. excel_path = excel_path
        self.timeout: int = timeout
        self.wb = load_workbook(excel_path)
        self.ws: Worksheet = self.wb.active

        self.payload = {
            'access_token': self.token,
            'v': '5.131'
        }

    def get_token(self, username, password):
        """
        Generate authentication token and user id for current session
        """
        resp: Response = requests.get(AUTH_URL + f'username={username}&password={password}')
        cont = json.loads(resp.text)
        token = cont['access_token']
        user_id = cont['user_id']
        print(resp.text)
        return token, user_id

    def send_message(self, user_id, message):
        """
        Send message to user by user_id. User id must be an integer or int32, something like that
        https://dev.vk.com/method/messages.send
        """
        rnd_i = random.randint(100, 100000000)
        payload = {'message': message, 'random_id': str(rnd_i), 'user_id': user_id}
        prm = self.payload | payload
        resp: Response = requests.request('GET', BASE_URL + SEND_MESSAGE_PATH, params=prm)
        return resp.text

    def get_user_id(self, user_key):
        """
        Converts custom or any user id to standard id.
        https://dev.vk.com/method/users.get
        Returns a string
        """
        payload = {'fields': 'bdate', 'user_ids': user_key}
        prm = self.payload | payload
        resp: Response = requests.request('GET', BASE_URL + GET_USER_PATH, params=prm)
        cont: dict = json.loads(resp.text)
        tmp_response: list = cont['response']
        u_id = tmp_response[0]['id']
        return str(u_id)

    def url_to_id(self, url: str):
        """
        Extracts user id from target url
        """
        x = url.split('/')
        usr_id: str = x[3]
        i = self.get_user_id(usr_id)
        return i

    def run(self):
        c1 = self.ws.cell(row=1, column=1)
        sheet = self.wb.worksheets[0]

        row_count = sheet.max_row
        column_count = sheet.max_column

        for r in range(row_count + 1):
            if r != 0 and r != 1:
                vkurl = self.ws[f'F{str(r)}'].value
                message = self.ws[f'G{str(r)}'].value
                mediaurl = self.ws[f'H{str(r)}'].value
                user_id = ms.url_to_id(url=vkurl)
                resp = ms.send_message(user_id=user_id, message=message)
                self.ws[f'J{str(r)}'].value = resp
                print(resp)
                time.sleep(self.timeout)
        self.wb.save(self.excel_path)


if __name__ == '__main__':
    ms = MessagingService(username='+380630597236', password='Ljn%26f%232!', excel_path='C:\\Users\\XYZ\\Documents\\VKDATA.xlsx', timeout=10)
    ms.run()
